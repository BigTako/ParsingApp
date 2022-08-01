from openpyxl import Workbook, load_workbook

FILEPATH = 'data.xlsx'
workbook = Workbook()
letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M',
           'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']

def insert_data(sheet:Workbook.active, start_indx:int, data:dict):
    """Insert data into sheet in pointed column
       sheet - Workbook().activate object
       data - contains dictionary with data(key will be column name in sheet)
    """
    letter_idx = 0
    for name, values in data.items():
        sheet[f'{letters[letter_idx]}1'] = name  # set column name(A1, B1...)
        for i in range(start_indx, start_indx + len(values)):
            sheet[letters[letter_idx]+str(i)] = values[i-start_indx]
        letter_idx += 1
    workbook.save(FILEPATH)  # save sheet into xlsx file (FILENAME)

def get_data_fro_col(worksheet:Workbook.active, col_name:str):
    """Function to get data from column by column name
        worksheet - Workbook().activate object
        col_name - name of column with needed data
    """
    values = []
    for col in worksheet.iter_cols(1, worksheet.max_column):
        if (col[0].value == col_name):
            for i in range(1, worksheet.max_row):
                values.append(col[i].value)
            break
    return values
